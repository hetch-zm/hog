"""Payroll core: Excel parsing, calculations, PDF payslip generation."""
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
)


REQUIRED_COLUMNS = [
    "Employee Number",
    "Full Name",
    "Position",
    "Pay Month",
    "Basic Salary",
    "Allowances",
    "Other Deductions",
    "Leave Days",
    "Bank Name",
    "Account Number",
    "Branch",
]

# Backward-compatible aliases: an old template with "Deductions" still parses.
COLUMN_ALIASES = {
    "Other Deductions": ["Other Deductions", "Deductions"],
}

# Calculated columns shown to the user as live previews.
CALC_COLUMNS = ["Gross Pay", "NAPSA", "NHIMA", "Taxable Income", "PAYE", "Net Pay"]


def _build_paye_formula(taxable_cell, bands):
    """Marginal-stack PAYE formula: SUM_i max(0, taxable - lower_i) * (rate_i - rate_{i-1}).
    Generalises to any monotonic band list. `bands` is a list of (lower, upper, rate) tuples."""
    sorted_bands = sorted(bands, key=lambda b: b[0])
    terms = []
    prev_rate = 0.0
    for lower, _upper, rate in sorted_bands:
        increment = rate - prev_rate
        if increment > 0:
            terms.append(f"MAX(0,{taxable_cell}-{lower})*{increment:.6f}")
        prev_rate = rate
    return "+".join(terms) if terms else "0"


def create_template(path, bands=None, settings=None, data_rows=20):
    """Generate a fresh template with formulas pre-filled in the calc columns.
    Bands and settings, if provided, are baked into the formulas as constants —
    so the Excel preview matches what the server will compute at upload time."""
    if bands is None:
        bands = [
            (0.0,    5100.0, 0.00),
            (5100.0, 7100.0, 0.20),
            (7100.0, 9200.0, 0.30),
            (9200.0, None,   0.37),
        ]
    if settings is None:
        settings = {}
    napsa_rate    = float(settings.get("napsa_rate", 0.05))
    napsa_ceiling = float(settings.get("napsa_ceiling", 28455.86))
    nhima_rate    = float(settings.get("nhima_rate", 0.01))
    napsa_ded = 1 if settings.get("napsa_deductible", True) else 0
    nhima_ded = 1 if settings.get("nhima_deductible", False) else 0

    all_columns = REQUIRED_COLUMNS + CALC_COLUMNS
    n_input = len(REQUIRED_COLUMNS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    ws.append(all_columns)

    hdr_input = PatternFill("solid", fgColor="0d3b66")
    hdr_calc = PatternFill("solid", fgColor="444444")
    calc_fill = PatternFill("solid", fgColor="eef1f7")
    totals_fill = PatternFill("solid", fgColor="d9e0eb")
    thin = Side(border_style="thin", color="cccccc")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for i, name in enumerate(all_columns, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True, color="ffffff")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = hdr_input if i <= n_input else hdr_calc
        cell.border = border

    samples = [
        ["EMP001", "Jane Doe",   "Accountant",   "April 2026",
         15000, 2500, 1800, 0, "ZANACO",  "1234567890", "Cairo Road"],
        ["EMP002", "John Smith", "Farm Manager", "April 2026",
         22000, 3500, 2700, 2, "Stanbic", "9876543210", "Manda Hill"],
    ]

    for i in range(data_rows):
        r = i + 2
        if i < len(samples):
            for j, v in enumerate(samples[i], start=1):
                ws.cell(row=r, column=j, value=v)

        emp = f"A{r}"
        basic, allow, other = f"E{r}", f"F{r}", f"G{r}"
        gross, napsa, nhima, taxable, paye, net = (
            f"L{r}", f"M{r}", f"N{r}", f"O{r}", f"P{r}", f"Q{r}"
        )

        ws[gross]   = f'=IF({emp}="","",{basic}+{allow})'
        ws[napsa]   = f'=IF({gross}="","",MIN({gross},{napsa_ceiling})*{napsa_rate})'
        ws[nhima]   = f'=IF({gross}="","",{gross}*{nhima_rate})'
        ws[taxable] = f'=IF({gross}="","",{gross}-{napsa}*{napsa_ded}-{nhima}*{nhima_ded})'
        paye_expr = _build_paye_formula(taxable, bands)
        ws[paye]    = f'=IF({taxable}="","",{paye_expr})'
        ws[net]     = f'=IF({gross}="","",{gross}-{napsa}-{nhima}-{paye}-{other})'

    # Totals row.
    totals_row = data_rows + 2
    ws.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True)
    sum_cols = ["E", "F", "G", "L", "M", "N", "O", "P", "Q"]
    for col in sum_cols:
        col_idx = ord(col) - ord("A") + 1
        rng = f"{col}2:{col}{data_rows + 1}"
        cell = ws.cell(row=totals_row, column=col_idx, value=f"=SUM({rng})")
        cell.font = Font(bold=True)

    money_cols = ["E", "F", "G", "L", "M", "N", "O", "P", "Q"]
    for col in money_cols:
        for r in range(2, totals_row + 1):
            ws[f"{col}{r}"].number_format = '#,##0.00'

    for col in ["L", "M", "N", "O", "P", "Q"]:
        for r in range(2, data_rows + 2):
            ws[f"{col}{r}"].fill = calc_fill
    for col in sum_cols + ["A"]:
        ws[f"{col}{totals_row}"].fill = totals_fill
        ws[f"{col}{totals_row}"].border = border

    for i, name in enumerate(all_columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(13, len(name) + 2)
    ws.freeze_panes = "A2"

    # Note row about formula provenance.
    note_row = totals_row + 2
    note = ws.cell(
        row=note_row, column=1,
        value=(
            "Tip: The grey columns (Gross, NAPSA, NHIMA, Taxable, PAYE, Net) are computed by "
            "Excel formulas using the rates set by the administrator at the time you downloaded "
            "this template. The server recomputes these at upload time using whatever the rates "
            "are then — your final payslip values are always the server's values."
        ),
    )
    note.font = Font(italic=True, color="666666")
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=len(all_columns))
    ws.row_dimensions[note_row].height = 36

    wb.save(path)


def _to_float(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _resolve_column(headers, name):
    candidates = COLUMN_ALIASES.get(name, [name])
    for c in candidates:
        if c in headers:
            return headers.index(c)
    return None


def parse_excel(path):
    """Read the uploaded Excel and return a list of employee dicts.
    Net pay is recomputed server-side after statutory; what we return here
    treats `deductions` as 'other deductions only'."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty")

    headers = [str(c).strip() if c else "" for c in rows[0]]
    idx = {}
    missing = []
    for col in REQUIRED_COLUMNS:
        i = _resolve_column(headers, col)
        if i is None:
            missing.append(col)
        else:
            idx[col] = i
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")

    employees = []
    for row in rows[1:]:
        if not row or all(c is None or c == "" for c in row):
            continue
        emp_no = row[idx["Employee Number"]]
        if not emp_no:
            continue
        basic = _to_float(row[idx["Basic Salary"]])
        allow = _to_float(row[idx["Allowances"]])
        deduc = _to_float(row[idx["Other Deductions"]])
        leave_days = _to_float(row[idx["Leave Days"]])
        gross = basic + allow
        employees.append({
            "employee_no": str(emp_no).strip(),
            "full_name": str(row[idx["Full Name"]] or "").strip(),
            "position": str(row[idx["Position"]] or "").strip(),
            "pay_month": str(row[idx["Pay Month"]] or "").strip(),
            "basic_salary": basic,
            "allowances": allow,
            "deductions": deduc,
            "leave_days": leave_days,
            "bank_name": str(row[idx["Bank Name"]] or "").strip(),
            "account_number": str(row[idx["Account Number"]] or "").strip(),
            "branch": str(row[idx["Branch"]] or "").strip(),
            "gross_pay": gross,
            "net_pay": gross - deduc,  # placeholder; app.py recomputes after statutory
        })
    if not employees:
        raise ValueError("No employee rows found")
    return employees


def _money(v):
    return f"{v:,.2f}"


def generate_payslip_pdf(emp, batch_meta, out_path,
                         company_name="Hope of Glory",
                         currency="ZMW",
                         ytd=None):
    doc = SimpleDocTemplate(
        out_path, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title", parent=styles["Heading1"],
        fontSize=18, alignment=1, spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=10, alignment=1, textColor=colors.grey, spaceAfter=14,
    )
    section_style = ParagraphStyle(
        "section", parent=styles["Heading3"],
        fontSize=11, spaceBefore=4, spaceAfter=4,
    )

    story = [
        Paragraph(company_name, title_style),
        Paragraph(f"Payslip — {emp['pay_month']}", sub_style),
    ]

    info = [
        ["Employee Number", emp["employee_no"], "Pay Month", emp["pay_month"]],
        ["Full Name", emp["full_name"], "Position", emp["position"]],
        ["Bank", emp["bank_name"], "Account No.", emp["account_number"]],
        ["Branch", emp["branch"], "Leave Days", _money(emp.get("leave_days", 0))],
    ]
    info_tbl = Table(info, colWidths=[32 * mm, 55 * mm, 32 * mm, 55 * mm])
    info_tbl.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.4, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("BACKGROUND", (2, 0), (2, -1), colors.whitesmoke),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 8 * mm))

    napsa = emp.get("napsa", 0) or 0
    nhima = emp.get("nhima", 0) or 0
    paye = emp.get("paye", 0) or 0
    other = emp.get("deductions", 0) or 0
    total_deductions = napsa + nhima + paye + other

    breakdown = [
        [f"Earnings", f"Amount ({currency})", "Deductions", f"Amount ({currency})"],
        ["Basic Salary", _money(emp["basic_salary"]),  "NAPSA",            _money(napsa)],
        ["Allowances",  _money(emp["allowances"]),     "NHIMA",            _money(nhima)],
        ["",            "",                            "PAYE",             _money(paye)],
        ["",            "",                            "Other Deductions", _money(other)],
        ["Gross Pay",   _money(emp["gross_pay"]),      "Total Deductions", _money(total_deductions)],
    ]
    bd_tbl = Table(breakdown, colWidths=[44 * mm, 44 * mm, 44 * mm, 42 * mm])
    bd_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d3b66")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eef1f7")),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("ALIGN", (3, 1), (3, -1), "RIGHT"),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(bd_tbl)
    story.append(Spacer(1, 6 * mm))

    net_tbl = Table(
        [["NET PAY", f"{currency}  {_money(emp['net_pay'])}"]],
        colWidths=[88 * mm, 86 * mm],
    )
    net_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0d3b66")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 13),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(net_tbl)
    story.append(Spacer(1, 8 * mm))

    if ytd:
        story.append(Paragraph(f"Year to date ({ytd.get('year', '')})", section_style))
        ytd_tbl = Table([
            ["Gross", "NAPSA", "NHIMA", "PAYE", "Other", "Net"],
            [
                _money(ytd.get("gross", 0)),
                _money(ytd.get("napsa", 0)),
                _money(ytd.get("nhima", 0)),
                _money(ytd.get("paye", 0)),
                _money(ytd.get("deductions", 0)),
                _money(ytd.get("net", 0)),
            ],
        ], colWidths=[29 * mm] * 6)
        ytd_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#444")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 1), (-1, 1), "RIGHT"),
            ("BOX", (0, 0), (-1, -1), 0.4, colors.grey),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(ytd_tbl)
        story.append(Spacer(1, 8 * mm))

    approvals = [
        ["Initiated by", batch_meta.get("initiator", "-"), batch_meta.get("initiated_at", "")],
        ["Approver 1", batch_meta.get("approver1", "-"), batch_meta.get("approver1_at", "")],
        ["Approver 2", batch_meta.get("approver2", "-"), batch_meta.get("approver2_at", "")],
    ]
    ap_tbl = Table(
        [["Approval Trail", "Name", "Timestamp"]] + approvals,
        colWidths=[40 * mm, 70 * mm, 64 * mm],
    )
    ap_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#444")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(ap_tbl)

    footer = ParagraphStyle(
        "footer", parent=styles["Normal"],
        fontSize=8, alignment=1, textColor=colors.grey,
    )
    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph(
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} — This is a system-generated payslip.",
        footer,
    ))

    doc.build(story)
    return out_path
